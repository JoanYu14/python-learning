# openpyxl 是一個用於讀取和寫入 Excel 檔案的 Python 模組，支援 .xlsx 格式（Excel 2007 及以後版本）。它提供了簡單且靈活的方式來操作 Excel 工作簿，包括建立、讀取、修改、儲存電子表格內容。
# 由於Excel檔案（我們稱之為工作簿）可能包含多個工作表，因此為了從中讀取數據，我們需要：
#     1. 載入工作簿。
#     2. 取得具體工作表。
#     3. 迭代該工作表中的每一行。
#     4. 迭代該行中的每個單元格。
import os
from openpyxl import load_workbook

file_path = os.path.join(
    os.getcwd(), "14-ETL_Data_in_Python", "14_2-Excel", "Dodgers.xlsx"
)

# 1. Load the workbook載入工作簿
# 使用 load_workbook 函數加載指定路徑的 Excel 文件，並將其賦值給 wb 變數
# wb 是一個 openpyxl 的工作簿對象

wb = load_workbook(file_path)

# 打印工作簿對象，確認加載成功
print(wb)

# 用於存儲工作表中每一行的數據
result = []

# 2. load worksheet載入工作表
# 從工作簿中載入第一個工作表 (索引 0 表示第一個工作表)
ws = wb.worksheets[0]

# 使用 iter_rows() 方法遍歷工作表中的每一行
for row in ws.iter_rows():
    # 將每一行中每個單元格的值提取出來，存入 result 列表
    # 使用列表推導式
    # for cell in row:
    #  這是 for 循環的一部分，它逐個遍歷 row 中的每個 cell（單元格）。row 是由 ws.iter_rows() 返回的行，其中每一行包含一個由單元格對象組成的元組。
    #  cell 代表每個單元格對象，從該單元格中，我們可以提取其屬性和方法。
    # cell.value:代表每個單元格中的值。
    # [cell.value for cell in row]
    #   這是 列表推導式，它是一種簡潔的方式來創建列表。這個語法會創建一個列表，列表中的每個元素是 row 中每個 cell 的 .value。
    result.append([cell.value for cell in row])

# print(result)


# 用來計算全壘打總數的變量
sum = 0

# 遍歷 result 中的數據，從第二行開始（跳過表頭）
for r in result[1:]:
    # r[11] 表示該行的第 12 列（索引從 0 開始），將其轉換為整數並加到 sum 中
    sum += int(r[11])

print(f"道奇隊的全壘打總數為{sum}")  # 道奇隊的全壘打總數為168
